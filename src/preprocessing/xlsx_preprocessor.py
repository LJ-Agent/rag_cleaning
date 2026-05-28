"""Excel (XLSX) preprocessor — extract all sheets as pages, parse cells/merged cells/formulas."""

import io
from uuid import uuid4

from common.config_loader import get_config
from common.exception.exceptions import PreprocessingException
from common.models.document import (
    Document,
    DocumentMetadata,
    ElementRole,
    Page,
    TableCell,
    TableElement,
    TextElement,
)
from common.util.logger import bind_trace_id, get_logger
from common.util.utils import md5_bytes
from preprocessing.base import BasePreprocessor

logger = get_logger()


class XlsxPreprocessor(BasePreprocessor):
    """Extract raw cell data from Excel workbooks. Each sheet becomes a page."""

    format_type = "xlsx"
    supported_extensions = {"xlsx", "xls"}
    supported_mime_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }

    def __init__(self):
        cfg = get_config()["preprocessing"]["xlsx"]
        self._max_rows = cfg.get("max_rows", 100000)
        self._max_columns = cfg.get("max_columns", 500)
        self._merge_strategy = cfg.get("merge_cell_strategy", "fill")

    def extract(self, file_data: bytes, file_name: str) -> Document:
        log = bind_trace_id(str(uuid4())[:8])
        doc_id = md5_bytes(file_data)[:16]
        doc = Document(doc_id=doc_id)

        doc.metadata = DocumentMetadata(
            source_format="xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_size_bytes=len(file_data),
            file_md5=md5_bytes(file_data),
        )
        doc.log_stage("xlsx_preprocess_start")

        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(io.BytesIO(file_data), data_only=True, read_only=True)

            # Metadata from workbook properties
            if wb.properties:
                doc.metadata.title = wb.properties.title or ""
                doc.metadata.author = wb.properties.creator or ""
                doc.metadata.created_at = str(wb.properties.created) if wb.properties.created else ""

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                page = Page(page_number=len(doc.pages) + 1)

                # Build table from worksheet data
                rows = []
                row_count = 0
                full_text_parts = []

                for row in ws.iter_rows(
                    max_row=min(ws.max_row or 0, self._max_rows),
                    max_col=min(ws.max_column or 0, self._max_columns),
                    values_only=False,
                ):
                    cells = []
                    for cell in row:
                        value = cell.value
                        text = str(value) if value is not None else ""
                        full_text_parts.append(text)
                        cells.append(TableCell(text=text))
                    if any(c.text for c in cells):  # Skip completely empty rows
                        rows.append(cells)
                        row_count += 1

                doc.metadata.char_count += sum(len(c.text) for c in sum(rows, []))
                doc.metadata.word_count += len(" ".join(full_text_parts).split())
                doc.metadata.has_tables = True

                # Add table element for the sheet
                table = TableElement(
                    element_id=self._make_element_id(doc_id, page.page_number, 0, "tbl"),
                    role=ElementRole.TABLE,
                    page_numbers=[page.page_number],
                    rows=rows,
                    column_count=max((len(r) for r in rows), default=0),
                    row_count=len(rows),
                )
                page.elements.append(table)
                page.text_content = "\n".join(full_text_parts)

                # Add sheet name as heading
                heading = TextElement(
                    element_id=self._make_element_id(doc_id, page.page_number, 1, "txt"),
                    role=ElementRole.HEADING,
                    page_numbers=[page.page_number],
                    text=f"## Sheet: {sheet_name}",
                )
                page.elements.insert(0, heading)

                doc.pages.append(page)

            wb.close()

        except Exception as e:
            raise PreprocessingException(f"XLSX extraction failed: {e}", format_type="xlsx")

        doc.metadata.page_count = len(doc.pages)
        doc.log_stage("xlsx_preprocess_done")
        return doc
